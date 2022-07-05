import requests
from bs4 import BeautifulSoup
import os
import json
import sys
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages")
from requests_oauthlib import OAuth1Session
sys.path.append("/Users/kuramochiosuke/.pyenv/versions/3.10.4/lib/python3.10/site-packages")
from dotenv import find_dotenv, load_dotenv
import requests
import schedule
import time
import openpyxl


def job():
    # .envファイルを探して読み込み
    env_file = find_dotenv()
    load_dotenv(env_file)  

    CONSUMER_KEY = os.environ.get('CONSUMER_KEY')
    CONSUMER_SECRET = os.environ.get('CONSUMER_SECRET')
    ACCESS_KEY = os.environ.get('ACCESS_KEY')
    ACCESS_KEY_SECRET = os.environ.get('ACCESS_KEY_SECRET')

    # Twitterの認証
    twitter = OAuth1Session(CONSUMER_KEY, CONSUMER_SECRET, ACCESS_KEY, ACCESS_KEY_SECRET)
    print(twitter)

    #エンドポイント
    url_text = 'https://api.twitter.com/1.1/statuses/update.json'
    url_media = "https://upload.twitter.com/1.1/media/upload.json"



    # ここまでTwitter投稿の準備
    print("投稿準備完了")
    for i in range(12):
        print(i)
        try:
            #スニだんのページの指定
            URL = 'https://snkrdunk.com'
            # リクエストヘッダの指定
            headers = {"User-Agent": "hoge"}
            response = requests.get(URL,  headers=headers)
            r_text=response.text
            soup = BeautifulSoup(r_text, 'html.parser')


            print('スニダンのページの取得完了')



            # 記事の取得 
            soup_article=soup.find_all("article",attrs={"class","article-list new"})[i]      
            soup_text=soup_article.find_all("h3")[0].find("a").text.replace("\n","").replace("\t","").replace("定価/","")
            # 画像の取得
            soup_img=soup_article.find_all("img")[0]['src']
            # 詳細ページのリンクを取得
            soup_url="https://snkrdunk.com/"+soup_article.find("a")['href']
            # 画像の処理
            response = requests.get(soup_img)
            image = response.content
            files = {"media" : image}
            req_media = twitter.post(url_media, files = files)
            media_id = json.loads(req_media.text)['media_id']
            print('画像の取得完了')


            print("ここから詳細ページ")
            
            URL = soup_url
            # リクエストヘッダの指定
            headers = {"User-Agent": "hoge"}
            response = requests.get(URL,  headers=headers)
            r_text=response.text
            soup = BeautifulSoup(r_text, 'html.parser')

            if soup_text[:5]=="【リーク】":
                soup_h1=soup.find_all("h1",attrs={"class","page-title"})[0].text.replace("抽選/定価/販売店舗まとめ","").replace("【リーク】","")
                soup_text=soup.find_all("div",attrs={"class","article-content"})[0].text.replace("\n","").replace("\t","")
                pos1=soup_text.find("について")
                text1=soup_text[pos1+4:]
                pos2=text1.find("発売予定！")
                if pos2<10: 
                    pos3=text1.find("発売！")
                    if pos3<10:
                        pos4=text1.find("復刻予定！")
                        if pos4<10:
                            pos5=text1.find("リリース！")
                            if pos5<10:
                                pos6=text1.find("リリース予定！")
                                soup_cap=text1[:pos6+7]
                                params = {'status':"リーク情報!!!\n\n{}\n\n{}\n\n情報が入り次第更新!!!".format(soup_h1,soup_cap),'media_ids':[media_id]}
                            else:
                                soup_cap=text1[:pos5+5]
                                params = {'status':"リーク情報!!!\n\n{}\n\n{}\n\n情報が入り次第更新!!!".format(soup_h1,soup_cap),'media_ids':[media_id]}
                        else:
                            soup_cap=text1[:pos4+5]
                            params = {'status':"リーク情報!!!\n\n{}\n\n{}\n\n情報が入り次第更新!!!".format(soup_h1,soup_cap),'media_ids':[media_id]}
                    else:
                        soup_cap=text1[:pos3+3]
                        params = {'status':"リーク情報!!!\n\n{}\n\n{}\n\n情報が入り次第更新!!!".format(soup_h1,soup_cap),'media_ids':[media_id]}
                else:
                    soup_cap=text1[:pos2+5]
                    params = {'status':"リーク情報!!!\n\n{}\n\n{}\n\n情報が入り次第更新!!!".format(soup_h1,soup_cap),'media_ids':[media_id]}
                wb = openpyxl.load_workbook('sneaker.xlsx')
                ws = wb["Sheet1"]
                for i in range(wb['Sheet1'].max_row):
                    if ws.cell(row=i+1,column=1).value==params["status"]:
                        print("投稿済みです")
                        break    
                    elif i==wb['Sheet1'].max_row-1:  
                        ws.cell(row=wb['Sheet1'].max_row+1,column=1).value = params["status"]
                        wb.save('sneaker.xlsx')
                        print("保存しました")
                        twitter.post(url_text, params = params)
                        print("投稿しました")  
                print("")
            elif soup_text[:9]=="【販売リンクあり】":
                a_count=len(soup.find_all("div",attrs={"class","sneaker-release-shop-box pre-release"})[0].find_all("a"))
                for i in range(a_count):
                    soup_block=soup.find_all("div",attrs={"class","sneaker-release-shop-box"})[0].find_all("a")[i]
                    soup_link=soup_block['href']
                    soup_app_name=soup_block.find_all("div",attrs={"class","left-box"})[0].text
                    soup_data=soup_block.find_all("div",attrs={"class","shop-right-box"})[0].text
                    params = {'status': "{}\n\n{}  {}\n{}\n".format(soup_text,soup_app_name,soup_data,soup_link),'media_ids':[media_id]}
                    wb = openpyxl.load_workbook('sneaker.xlsx')
                    ws = wb["Sheet1"]
                    for i in range(wb['Sheet1'].max_row):
                        if ws.cell(row=i+1,column=1).value==params["status"]:
                            print("投稿済みです")
                            break    
                        elif i==wb['Sheet1'].max_row-1:   
                            ws.cell(row=wb['Sheet1'].max_row+1,column=1).value = params["status"]
                            wb.save('sneaker.xlsx')
                            print("保存しました")
                            twitter.post(url_text, params = params)
                            print("投稿しました")  
                    print("")
            else:
                print("除外")
                print("")
        except IndexError:
            print("INDEX エラーです")       
            print("")
        except FileNotFoundError:
            print("NOT FILE エラーです")       
            print("")
        except KeyError:
            print("KeyError エラーです")       
            print("")
    print("処理終了")


                
def main():
    schedule.every(1).minutes.do(job)
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == '__main__':
    main()